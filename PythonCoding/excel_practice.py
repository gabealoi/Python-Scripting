# IMPORTS
from ast import Str
from cgi import print_form
from tokenize import String
import openpyxl
# END IMPORTS

alphabet = {1:'a', 2:'b', 3:'c', 4:'d', 5:'e', 6:'f', 7:'g', 8:'h', 9:'i', 10:'j', 11:'k', 12:'l', 13:'m', 14:'n', 15:'o', 16:'p', 17:'q', 18:'r', 19:'s', 20:'t', 21:'u', 22:'v', 23:'w', 24:'x', 25:'y', 26:'z'} 

def main():
    f = "emp_id.xlsx"
    choice = -1

    while choice not in range(1,4):
        print("~~~Menu~~~")
        print("1 => Update Excel")
        print("2 => Read Excel")
        print("3 => Exit")
        print("\n")

        try:
            choice = int(input("Please Enter the Number Representing Your Choice: "))
        except ValueError:
            choice = -1
            print("<INVALID INPUT>, please enter a number representing your selection\n")

    if choice == 1:
        try:
            createFile(f)
        except PermissionError:
            print("Please close the file before running this program")
    elif choice == 2:
        readFile(f)
    elif choice == 3:
        print("Exiting..")
        exit()
    

def readFile(f):
    noError = True
    # load excel file
    try:
        workbook = openpyxl.load_workbook(filename=f)
    except PermissionError:
        noError = False
        print("Please close the file before running this program")
    except FileNotFoundError or FileExistsError:
        noError = False
        print("The file was not found or does not exist")

    if noError:
        empToName = {}
        sheet = workbook.active
        rows = sheet.max_row

        for i in range(2, rows+1):
            idCells = sheet.cell(row=i, column=1)
            nameCells = sheet.cell(row=i, column=2)
            empID = idCells.value
            empName = nameCells.value
            # update dictionary
            empToName[empID] = empName

            # print(str(idCells.value).zfill(4))
        for k in empToName:
            print("%-6d%-2s%-30s" % (k, "|", empToName.get(k)))
        
        print("\n")
        main()


def createFile(f):
    noError = True

    # load excel file
    try:
        workbook = openpyxl.load_workbook(filename=f)
    except PermissionError:
        noError = False
        print("Please close the file before running this program")
    except FileNotFoundError or FileExistsError:
        noError = False
        print("The file was not found or does not exist")

        


    if noError:
        # open workbook
        sheet = workbook.active
        lowerCol = 1
        upperCol = sheet.max_column
        colSelection = "~"
        rowSelection = -1
        val = None

        while colSelection.lower() not in alphabet.values():
            print("You have chosen to update the file, please enter a col letter in the range of " + alphabet.get(1).capitalize() + " to " + alphabet.get(sheet.max_column).capitalize())
            colSelection = input()

        while rowSelection not in range(1, sheet.max_row):
            print("Please select a row between 1 and " + str(sheet.max_row))
            try:
                rowSelection = int(input())
            except ValueError:
                rowSelection = -1
                print("Please enter a valid number in the above-specified range")

        if colSelection.lower() == "a":
            while val == None or len(str(val)) > 4:
                if colSelection.capitalize() == "A":
                    print("Please enter a new 4-digit employee ID that is not a current ID for: " + sheet.cell(row=rowSelection, column=2).value)
                    try:
                        val = int(input())
                    except ValueError:
                        val = None
                        print("Please enter a valid 4-digit number for the employee ID")
        elif colSelection.lower() == "b":
            while val == None:
                print("Please enter a new name for employee with ID: " + str(sheet.cell(row=rowSelection, column=1).value))
                val = input()


        # modify value
        # sheet["A9"] = 654
        cell = colSelection + str(rowSelection)
        sheet[cell] = val

        # save the file
        workbook.save(filename="emp_id.xlsx")
        # workbook.save(filename="emp_id.xlsx")

        print("Data Written")
        main()

# main statement
if __name__ == "__main__":
    main()